from openpyxl import *

### This program finds the exact volume of air ventilation through the BIOPAC.###

# Given a row, this function will return the time stamp associated with the datum.
def time(row):
    return (row-2)/100

# Load raw data spreadsheet, organized by a title row and then following with data.
raw = load_workbook(filename = 'data.xlsx')
sheet = raw['Sheet1']

# Create list of columns to iterate through.
columns = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R']
dicts = []

# Collects all local extrema in data with the corresponding time stamp.
for i in columns:
    rowFirst = 2
    dictall = {}
    dictall[time(rowFirst)] = sheet[''.join([i,str(rowFirst)])].value

    while sheet[''.join([i,str(rowFirst+2)])].value:
        setV = [sheet[''.join([i,str(rowFirst)])].value,sheet[''.join([i,str(rowFirst+1)])].value,sheet[''.join([i,str(rowFirst+2)])].value]
        if setV[0] < setV[1] and setV[2] < setV[1]:
            dictall[time(rowFirst+1)] = setV[1]
        elif setV[0] > setV[1] and setV[2] > setV[1]:
            dictall[time(rowFirst+1)] = setV[1]
        elif setV[0] == setV[1]:
            dictall[time(rowFirst)] = setV[0]
        elif setV[2] == setV[1]:
            dictall[time(rowFirst+1)] = setV[1]
        rowFirst+=1
    dicts.append(dictall)
    raw.close()

# Writes all local extrema for each column into a separate spreadsheet.
araw = Workbook()
ws = araw.active
ws.title = "Sheet1"
for i in range(0,18):
    prev = [0,0]
    air = 0
    for j,k in enumerate(dicts[i].items()):
        ws.cell(row=j+1,column=(4*i)+1).value = k[0]
        ws.cell(row=j+1,column=(4*i)+2).value = k[1]
        ws.cell(row=j+1,column=(4*i)+3).value = k[0]-prev[0]
        ws.cell(row=j+1,column=(4*i)+4).value = k[1]-prev[1]
        prev = k
araw.save(filename = 'AnalyzeRaw.xlsx')

# ORIGINALLY TWO FILES: Reads the raw analyzed extremas and sums only the positive increases in air ventilation.
araw = load_workbook(filename = 'AnalyzeRaw.xlsx')
ws = araw['Sheet1']

# Calculates and prints out air ventilation in the first 15, 30, 45 seconds, and the total volume.
for i in range(0,18):
    rowA, air15, air30, air45,airtotal = 2, 0, 0, 0, 0
    while ws.cell(row=rowA,column=(4*i)+4).value:
        if ws.cell(row=rowA,column=(4*i)+1).value < 27 and ws.cell(row=rowA,column=(4*i)+4).value >= 0:
            air15+=ws.cell(row=rowA,column=(4*i)+4).value
        elif ws.cell(row=rowA,column=(4*i)+1).value < 56 and ws.cell(row=rowA,column=(4*i)+4).value >= 0:
            air30+=ws.cell(row=rowA,column=(4*i)+4).value
        elif ws.cell(row=rowA,column=(4*i)+1).value < 100 and ws.cell(row=rowA,column=(4*i)+4).value >= 0:
            air45+=ws.cell(row=rowA,column=(4*i)+4).value
        rowA+=1
    airtotal = air15+air30+air45

    print(air15)
    print(air30)
    print(air45)
    print(airtotal)

# The volumes can then be copied to a spreadsheet and the averages calculated to use in determining work and efficiency.
